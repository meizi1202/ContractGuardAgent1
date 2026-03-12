# -*- coding: utf-8 -*-
"""行业风险库加载器模块

提供行业风险规则的加载、管理和匹配功能
"""

import sys
import os
import re
import json
from typing import Dict, Any, List, Optional

# 模块级别编码配置 - 解决 Windows 控制台中文乱码问题
# 必须在任何其他代码之前执行
if sys.platform == 'win32':
    # 设置标准输出/错误编码为 UTF-8
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            # 如果无法重新配置，则设置环境变量
            os.environ['PYTHONIOENCODING'] = 'utf-8'
    if sys.stderr.encoding != 'utf-8':
        try:
            sys.stderr.reconfigure(encoding='utf-8')
        except Exception:
            pass


class industryRiskLoader:
    """行业风险规则加载器

    负责加载和管理各行业的风险规则，支持行业自动识别和风险匹配
    """

    def __init__(self, basePath: str = None):
        """初始化风险规则加载器

        Args:
            basePath: 风险规则文件的基础路径
        """
        if basePath is None:
            basePath = os.path.dirname(__file__)

        self.basePath = basePath
        self.rules = {}
        self.industryKeywords = {}
        self._loadAllRules()

    def _loadAllRules(self):
        """加载所有风险规则"""
        # 计算数据目录路径 - 知识库-风险规则库
        # 使用绝对路径计算
        currentFile = os.path.abspath(__file__)
        srcDir = os.path.dirname(currentFile)  # src 目录
        skillDir = os.path.dirname(srcDir)     # contract-reviewer 目录
        claudeDir = os.path.dirname(skillDir)  # .claude 目录
        projectRoot = os.path.dirname(claudeDir)  # 项目根目录

        # 如果项目根目录下没有 docs 目录，尝试从当前工作目录获取
        if not os.path.exists(os.path.join(projectRoot, 'docs')):
            projectRoot = os.getcwd()

        dataPath = os.path.join(projectRoot, 'docs', 'knowledge_base', 'risk_rules')

        # 尝试从 Excel 文件加载
        excelLoaded = self._loadFromExcel(dataPath)

        if excelLoaded:
            print(f"成功从 Excel 加载风险规则")
            return

        # 回退：从 Python 文件加载
        print("Excel 文件未找到，尝试从 Python 模块加载...")
        self._loadFromPython(srcPath)

    def _loadFromExcel(self, dataPath: str) -> bool:
        """从 Excel 文件加载风险规则

        Args:
            dataPath: 数据目录路径

        Returns:
            是否成功加载
        """
        # Excel 文件到行业 ID 的映射
        excelMapping = {
            "通用风险规则.xlsx": "general",
            "采购合同风险规则.xlsx": "procurement",
            "租赁合同风险规则.xlsx": "leasing",
            "服务合同风险规则.xlsx": "service",
        }

        # 行业名称映射
        industryNames = {
            "general": "通用风险规则",
            "procurement": "采购合同",
            "leasing": "租赁合同",
            "service": "服务合同",
        }

        # 行业关键词（从 Excel 内容中提取）
        industryKeywords = {
            "procurement": ["采购", "供货", "供应", "买卖", "订购", "供应商", "货物", "产品", "设备", "材料"],
            "leasing": ["租赁", "租用", "出租", "承租", "租金", "房东", "租户", "房屋租赁", "设备租赁", "场地租赁"],
            "service": ["服务", "算力", "技术服务", "运维", "云服务", "托管", "租赁", "租用"],
            "general": ["采购", "供货", "供应", "买卖", "订购", "供应商", "租赁", "租用", "出租", "承租"],
        }

        try:
            from openpyxl import load_workbook
        except ImportError:
            print("openpyxl 库未安装，无法从 Excel 加载")
            return False

        loaded = False

        for excelFile, industryId in excelMapping.items():
            filePath = os.path.join(dataPath, excelFile)
            if not os.path.exists(filePath):
                continue

            try:
                wb = load_workbook(filePath)
                ws = wb.active

                # 读取表头
                headers = [cell.value for cell in ws[1]]
                idIdx = headers.index("风险ID") if "风险ID" in headers else 0
                catIdx = headers.index("风险类别") if "风险类别" in headers else 1
                nameIdx = headers.index("风险名称") if "风险名称" in headers else 2
                levelIdx = headers.index("严重级别") if "严重级别" in headers else 3
                patternIdx = headers.index("匹配模式") if "匹配模式" in headers else 4
                suggestIdx = headers.index("建议") if "建议" in headers else 5
                legalIdx = headers.index("法律依据") if "法律依据" in headers else 6

                # 读取数据
                risks = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[nameIdx]:
                        continue

                    patternStr = str(row[patternIdx]) if row[patternIdx] else ""
                    patterns = patternStr.split("|") if patternStr else []
                    # 处理正则表达式中的反斜杠 - Excel会吃掉反斜杠，需要还原
                    patterns = [self._fixRegexBackslash(p.strip()) for p in patterns if p.strip()]

                    risks.append({
                        "id": str(row[idIdx]) if row[idIdx] else "",
                        "level": str(row[levelIdx]).lower() if row[levelIdx] else "warning",
                        "name": str(row[nameIdx]) if row[nameIdx] else "",
                        "patterns": patterns,
                        "suggestion": str(row[suggestIdx]) if row[suggestIdx] else "",
                        "legalReference": str(row[legalIdx]) if row[legalIdx] else None
                    })

                # 按类别分组
                categoryMap = {}
                for risk in risks:
                    cat = risk.get("category", risk.get("name", ""))
                    cat = str(row[catIdx]) if row[catIdx] else "其他风险"
                    risk["category"] = cat
                    if cat not in categoryMap:
                        categoryMap[cat] = []
                    categoryMap[cat].append(risk)

                # 构建规则数据结构
                riskCategories = []
                for catName, catRisks in categoryMap.items():
                    riskCategories.append({
                        "category": catName,
                        "name": catName,
                        "risks": catRisks
                    })

                self.rules[industryId] = {
                    "industry": industryId,
                    "name": industryNames.get(industryId, industryId),
                    "version": "1.0",
                    "riskCategories": riskCategories
                }
                self.industryKeywords[industryId] = industryKeywords.get(industryId, [])

                loaded = True
                print(f"已加载: {excelFile} ({len(risks)} 条规则)")

            except Exception as e:
                print(f"加载 {excelFile} 失败: {e}")

        return loaded

    def _loadFromPython(self, srcPath: str):
        """从 Python 模块加载风险规则（回退方案）

        Args:
            srcPath: 源代码目录路径
        """
        import sys

        if srcPath not in sys.path:
            sys.path.insert(0, srcPath)

        # 计算数据目录路径 - 知识库-风险规则库
        projectRoot = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        dataPath = os.path.join(projectRoot, 'docs', 'knowledge_base', 'risk_rules')
        if dataPath not in sys.path:
            sys.path.insert(0, dataPath)

        try:
            from base_risks import baseRisks
            from procurement import procurementRisks
            from leasing import leasingRisks
        except ImportError as e:
            print(f"导入失败: {e}")
            return

        self.rules['general'] = baseRisks
        self._extractKeywords(baseRisks)

        # 加载采购合同风险规则
        self.rules['procurement'] = procurementRisks
        self._extractKeywords(procurementRisks)

        # 加载租赁合同风险规则
        self.rules['leasing'] = leasingRisks
        self._extractKeywords(leasingRisks)

    def _extractKeywords(self, ruleData: Dict[str, Any]):
        """提取行业关键词用于自动识别

        Args:
            ruleData: 风险规则数据
        """
        industry = ruleData.get('industry', 'unknown')
        keywords = ruleData.get('industryKeywords', {})
        if keywords:
            self.industryKeywords[industry] = keywords
        else:
            # 从规则内容中提取关键词
            self.industryKeywords[industry] = {}

    def _fixRegexBackslash(self, pattern: str) -> str:
        """修复Excel中丢失的反斜杠

        Excel存储正则表达式时会吃掉反斜杠，此方法用于还原常见的正则表达式元字符

        Args:
            pattern: 从Excel读取的原始模式字符串

        Returns:
            修复后的模式字符串
        """
        # 常见的正则表达式元字符需要反斜杠
        metaChars = ['d', 'D', 'w', 'W', 's', 'S', 'b', 'B', 'n', 't', 'r']
        for char in metaChars:
            # 匹配如 \d 但没有反斜杠的情况
            # 例如将 "d%" 修复为 "\d%"
            if f'{char}%' in pattern and f'\\{char}%' not in pattern:
                pattern = pattern.replace(f'{char}%', f'\\{char}%')
            if f'{char}' in pattern and f'\\{char}' not in pattern:
                # 检查是否是单词边界等复杂情况
                pass
        return pattern

    def getRules(self, industry: str = None) -> Dict[str, Any]:
        """获取指定行业的风险规则

        Args:
            industry: 行业标识，None表示获取所有规则

        Returns:
            风险规则字典
        """
        if industry is None:
            return self.rules

        if industry in self.rules:
            return self.rules[industry]

        # 尝试模糊匹配
        for key in self.rules.keys():
            if industry.lower() in key.lower() or key.lower() in industry.lower():
                return self.rules[key]

        return self.rules.get('general', {})

    def getAllIndustries(self) -> List[Dict[str, Any]]:
        """获取所有支持的行业列表

        Returns:
            行业信息列表
        """
        result = []
        for key, rule in self.rules.items():
            result.append({
                'id': rule.get('industry', key),
                'name': rule.get('name', key),
                'version': rule.get('version', '1.0'),
                'description': rule.get('description', '')
            })
        return result

    def detectIndustry(self, content: str) -> List[Dict[str, Any]]:
        """自动识别合同所属行业

        Args:
            content: 合同文本内容

        Returns:
            匹配的行业列表，按匹配度排序
        """
        industryScores = {}

        # 计算每个行业的匹配分数
        for industry, keywordsData in self.industryKeywords.items():
            if not keywordsData:
                continue

            score = 0
            matchedKeywords = []

            # 支持两种格式：字典或列表
            if isinstance(keywordsData, dict):
                # 字典格式: {"类型": ["关键词1", "关键词2"]}
                for keywordType, keywords in keywordsData.items():
                    if isinstance(keywords, list):
                        for keyword in keywords:
                            if keyword in content:
                                score += 1
                                matchedKeywords.append(keyword)
            elif isinstance(keywordsData, list):
                # 列表格式: ["关键词1", "关键词2"]
                for keyword in keywordsData:
                    if keyword in content:
                        score += 1
                        matchedKeywords.append(keyword)

            if score > 0:
                industryScores[industry] = {
                    'score': score,
                    'matchedKeywords': matchedKeywords[:5]  # 最多返回5个
                }

        # 按分数排序
        sortedIndustries = sorted(
            industryScores.items(),
            key=lambda x: x[1]['score'],
            reverse=True
        )

        # 转换为结果格式
        result = []
        for industry, info in sortedIndustries:
            result.append({
                'industry': industry,
                'name': self.rules.get(industry, {}).get('name', industry),
                'score': info['score'],
                'matchedKeywords': info['matchedKeywords']
            })

        return result

    def assessRisks(self, content: str, industry: str = None) -> List[Dict[str, Any]]:
        """评估合同风险

        Args:
            content: 合同文本内容
            industry: 指定行业，None表示自动识别

        Returns:
            风险列表
        """
        risks = []

        # 自动识别行业
        if industry is None:
            detected = self.detectIndustry(content)
            if detected:
                # 优先使用非通用行业
                for ind in detected:
                    if ind['industry'] != 'general':
                        industry = ind['industry']
                        print(f"自动识别行业: {industry} (匹配度: {ind['score']})")
                        break
                else:
                    # 如果没有特定行业，使用通用
                    industry = detected[0]['industry']
                    print(f"自动识别行业: {industry}")

        # 获取要检查的规则
        rulesToCheck = []

        # 添加通用规则
        if 'general' in self.rules:
            rulesToCheck.append(self.rules['general'])

        # 添加指定行业规则
        if industry and industry in self.rules:
            rulesToCheck.append(self.rules[industry])

        # 执行风险匹配
        for ruleSet in rulesToCheck:
            riskCategories = ruleSet.get('riskCategories', [])
            for category in riskCategories:
                categoryName = category.get('name', '')
                categoryDesc = category.get('description', '')

                for risk in category.get('risks', []):
                    riskInfo = self._matchRisk(content, risk)
                    if riskInfo:
                        riskInfo['category'] = categoryName
                        riskInfo['categoryDescription'] = categoryDesc
                        riskInfo['industry'] = ruleSet.get('industry', 'general')
                        risks.append(riskInfo)

        # 去重（基于风险ID）
        seen = set()
        uniqueRisks = []
        for risk in risks:
            riskId = risk.get('id', '')
            if riskId not in seen:
                seen.add(riskId)
                uniqueRisks.append(risk)

        return uniqueRisks

    def _matchRisk(self, content: str, risk: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """匹配单个风险规则

        Args:
            content: 合同文本内容
            risk: 风险规则

        Returns:
            匹配结果，如果匹配成功返回风险信息，否则返回None
        """
        patterns = risk.get('patterns', [])
        if not patterns:
            return None

        # 尝试匹配每个模式
        for pattern in patterns:
            try:
                matches = re.findall(pattern, content, re.IGNORECASE)
                if matches:
                    matchedText = matches[0] if isinstance(matches[0], str) else str(matches[0])

                    return {
                        'id': risk.get('id', ''),
                        'level': risk.get('level', 'warning'),
                        'name': risk.get('name', ''),
                        'matched': matchedText,
                        'suggestion': risk.get('suggestion', ''),
                        'legalReference': risk.get('legalReference'),
                        'pattern': pattern
                    }
            except re.error:
                # 正则表达式错误，跳过
                continue

        return None

    def getRiskSummary(self, risks: List[Dict[str, Any]]) -> Dict[str, Any]:
        """生成风险摘要

        Args:
            risks: 风险列表

        Returns:
            风险摘要
        """
        critical = sum(1 for r in risks if r.get('level') == 'critical')
        warning = sum(1 for r in risks if r.get('level') == 'warning')
        info = sum(1 for r in risks if r.get('level') == 'info')

        # 按行业分组
        byIndustry = {}
        for risk in risks:
            industry = risk.get('industry', 'unknown')
            if industry not in byIndustry:
                byIndustry[industry] = []
            byIndustry[industry].append(risk)

        return {
            'total': len(risks),
            'critical': critical,
            'warning': warning,
            'info': info,
            'byIndustry': byIndustry
        }


# 全局实例
_riskLoader = None


def getRiskLoader() -> industryRiskLoader:
    """获取全局风险加载器实例（单例模式）"""
    global _riskLoader
    if _riskLoader is None:
        _riskLoader = industryRiskLoader()
    return _riskLoader
