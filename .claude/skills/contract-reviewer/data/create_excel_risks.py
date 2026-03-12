# -*- coding: utf-8 -*-
"""创建行业风险库的 Excel 示例文件

将风险规则导出为 Excel 格式，方便非技术人员维护
"""

import sys
import os

# 模块级别编码配置 - 解决 Windows 控制台中文乱码问题
# 必须在任何其他代码之前执行
if sys.platform == 'win32':
    # 设置标准输出/错误编码为 UTF-8
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            # 如果无法重新配置，则设置环境变量
            os.environ['PYTHONIOENCODING'] = 'utf-8'
    if sys.stderr.encoding != 'utf-8':
        try:
            sys.stderr.reconfigure(encoding='utf-8')
        except Exception:
            pass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建输出目录
outputDir = os.path.join(os.path.dirname(__file__), 'data', 'industry_risks')
os.makedirs(outputDir, exist_ok=True)

# 定义表头样式
headerFont = Font(bold=True, color="FFFFFF")
headerFill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thinBorder = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def createBaseRisksExcel():
    """创建通用风险规则 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "通用风险规则"

    # 表头
    headers = ["风险ID", "风险类别", "风险名称", "严重级别", "匹配模式", "建议", "法律依据"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = headerFont
        cell.fill = headerFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thinBorder

    # 数据
    data = [
        ["AMT001", "金额风险", "合同金额未明确", "critical", "金额.*未确定|金额.*待定|金额.*另行约定", "建议明确约定具体金额", "《民法典》第543条"],
        ["AMT002", "金额风险", "金额大小写不一致", "warning", "(\\d+)元.*[零壹贰叁肆伍陆柒捌玖拾佰仟萬]+", "核对大小写金额是否一致", ""],
        ["TRM001", "期限风险", "合同期限过长", "warning", "期限.*[1-9]\\d{1,2}年|永久|无固定期限", "建议设置合理期限，并约定续期条件", ""],
        ["TRM002", "期限风险", "期限约定不明确", "warning", "期限.*另行通知|有效期.*待定|开始时间.*待定", "明确约定合同起止时间", ""],
        ["LIA001", "责任风险", "单方面免责条款", "warning", "免责|不承担.*责任|概不负责|不负.*任何责任", "注意免责条款的合理性，双方责任应对等", "《民法典》第497条"],
        ["LIA002", "责任风险", "违约金比例过高", "critical", "违约金.*[3-9]\\d%|违约金.*1\\d\\d%|日息.*千分之[5-9]", "违约金过高的，可能不受法律保护", "《民法典》第585条"],
        ["TRM003", "终止风险", "单方面解除限制", "warning", "不得解除|不可终止|单方面.*解除.*无效", "注意解约条款的公平性，保障双方权利", "《民法典》第562条"],
        ["TRM004", "终止风险", "终止条件不对等", "warning", "甲方.*有权.*终止|乙方.*无权.*终止", "确保双方终止权利对等", ""],
        ["DSP001", "争议解决风险", "争议解决条款缺失", "info", "争议.*未约定|纠纷.*另行协商", "建议补充争议解决条款", ""],
        ["DSP002", "争议解决风险", "管辖约定不明", "info", "管辖.*待定|法院.*另行约定", "明确约定管辖法院", "《民事诉讼法》第34条"],
    ]

    for row in data:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20

    wb.save(os.path.join(outputDir, "通用风险规则.xlsx"))
    print(f"已创建: 通用风险规则.xlsx")


def createProcurementRisksExcel():
    """创建采购合同风险规则 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "采购合同风险规则"

    # 表头
    headers = ["风险ID", "风险类别", "风险名称", "严重级别", "匹配模式", "建议", "法律依据"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = headerFont
        cell.fill = headerFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thinBorder

    # 数据
    data = [
        ["PUR_PAY001", "付款风险", "付款条件不明确", "critical", "付款方式.*待定|付款时间.*另行通知|付款.*另行协商", "明确约定付款方式、付款时间和付款比例", "《民法典》第628条"],
        ["PUR_PAY002", "付款风险", "预付款比例过高", "critical", "预付款.*[7-9]\\d%|预付款.*100%", "预付款比例过高会增加采购方风险，建议控制在30%以内", ""],
        ["PUR_PAY003", "付款风险", "验收后付款延迟", "warning", "验收.*\\d+.*日后.*付款|验收后.*个月.*付款", "明确验收标准和付款时限", ""],
        ["PUR_PAY004", "付款风险", "付款条件过于苛刻", "warning", "付款.*须.*完全.*满意|付款.*须.*甲方.*书面确认", "付款条件应具体可执行", ""],
        ["PUR_DEL001", "交付风险", "交付标准不清", "critical", "交付标准.*待定|交付.*另行约定|货物.*规格.*未明确", "明确货物的规格、型号、数量、质量标准", "《民法典》第511条"],
        ["PUR_DEL002", "交付风险", "交货地点不明确", "warning", "交货地点.*待定|交货地点.*另行通知|交货.*地点.*未约定", "明确交货地点及运输费用承担", ""],
        ["PUR_DEL003", "交付风险", "验收流程缺失", "warning", "验收.*未约定|验收.*另行协商|无需.*验收", "明确验收标准、流程和时限", "《民法典》第620条"],
        ["PUR_DEL004", "交付风险", "交货时间不明确", "warning", "交货时间.*待定|交货.*另行通知|交付时间.*未约定", "明确具体交货时间和分期交货安排", ""],
        ["PUR_QLT001", "质量风险", "质量标准模糊", "critical", "质量.*符合.*相关标准|质量.*达到.*行业标准|质量.*按照.*国家标准", "明确具体质量标准，最好有量化指标", "《民法典》第509条"],
        ["PUR_QLT002", "质量风险", "保修期不合理", "warning", "保修期.*少于.*\\d+个月|无保修|不享受.*保修", "根据产品特性约定合理保修期", "《产品质量法》第40条"],
        ["PUR_QLT003", "质量风险", "检验方法不明确", "warning", "检验方法.*另行约定|以.*甲方.*检验.*为准", "明确检验方法和标准", "《民法典》第620条"],
        ["PUR_QLT004", "质量风险", "质量异议期限过短", "warning", "异议期.*\\d+天|质量.*问题.*\\d+天内.*提出", "质量异议期限应合理", ""],
        ["PUR_PEN001", "违约风险", "违约金比例失衡", "critical", "甲方.*违约金.*\\d+%.+乙方.*违约金|乙方.*违约金.*比例.*高于.*甲方", "双方违约金比例应大致对等", "《民法典》第585条"],
        ["PUR_PEN002", "违约风险", "逾期交货责任过轻", "warning", "逾期交货.*仅.*支付.*\\d+%|逾期.*赔偿.*不超过.*合同金额", "逾期交货责任应与采购方损失相当", ""],
        ["PUR_PEN003", "违约风险", "单方面终止责任不对等", "warning", "甲方.*终止.*须.*赔偿|乙方.*终止.*不.*承担.*责任", "确保双方终止合同的责任对等", ""],
        ["PUR_TTL001", "所有权风险", "所有权转移时间不明确", "warning", "所有权.*另行约定|所有权.*转移.*时间.*未约定", "明确所有权转移时间和条件", "《民法典》第224条"],
        ["PUR_TTL002", "所有权风险", "知识产权归属不清", "info", "知识产权.*另行协商|技术.*成果.*归属.*未明确", "明确知识产权归属", ""],
    ]

    for row in data:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20

    wb.save(os.path.join(outputDir, "采购合同风险规则.xlsx"))
    print(f"已创建: 采购合同风险规则.xlsx")


def createLeasingRisksExcel():
    """创建租赁合同风险规则 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "租赁合同风险规则"

    # 表头
    headers = ["风险ID", "风险类别", "风险名称", "严重级别", "匹配模式", "建议", "法律依据"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = headerFont
        cell.fill = headerFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thinBorder

    # 数据
    data = [
        ["LEA_RNT001", "租金风险", "租金约定不明确", "critical", "租金.*另行协商|租金.*待定|租金.*未约定", "明确约定租金金额、支付方式和支付时间", "《民法典》第704条"],
        ["LEA_RNT002", "租金风险", "租金调整机制缺失", "warning", "租金.*不可调整|租金.*不再变动|租金.*保持不变", "建议约定租金调整机制，应对市场变化", ""],
        ["LEA_RNT003", "租金风险", "押金退还条件不清", "warning", "押金.*不予退还|押金.*退还.*另行约定|押金.*退还.*条件.*未明确", "明确押金退还条件和时限", "《民法典》第721条"],
        ["LEA_RNT004", "租金风险", "逾期滞纳金过高", "warning", "滞纳金.*日.*千分之|逾期.*罚款.*\\d+%|滞纳金.*超过.*本金", "滞纳金过高的可能不受法律保护", "《民法典》第585条"],
        ["LEA_USG001", "使用风险", "用途限制过严", "warning", "仅限.*使用|不得.*改变.*用途|擅自.*用途.*违约", "用途限制应合理", ""],
        ["LEA_USG002", "使用风险", "转租限制不合理", "critical", "不得.*转租|禁止.*转租.*任何.*情况|转租.*须.*支付.*高额.*违约金", "应允许在一定条件下转租", "《民法典》第716条"],
        ["LEA_USG003", "使用风险", "装修限制过严", "warning", "不得.*装修|禁止.*改造|装修.*须.*支付.*高额.*费用", "装修限制应合理", "《民法典》第715条"],
        ["LEA_USG004", "使用风险", "日常维护责任不明", "info", "维护.*责任.*未约定|维修.*另行协商", "明确日常维护责任划分", "《民法典》第712条"],
        ["LEA_MTN001", "维修风险", "维修责任不清", "warning", "维修.*责任.*未约定|维修.*另行协商|所有.*维修.*乙方.*承担", "明确出租人和承租人的维修责任", "《民法典》第712条"],
        ["LEA_MTN002", "维修风险", "维护标准缺失", "warning", "维护标准.*未约定|维护要求.*另行通知", "明确租赁物的维护标准和要求", ""],
        ["LEA_MTN003", "维修风险", "维修响应时间过长", "warning", "维修.*\\d+.*工作日|维修.*\\d+.*天后.*响应", "维修响应时间应合理", ""],
        ["LEA_TER001", "终止风险", "提前解约条件不对等", "critical", "甲方.*有权.*随时.*终止|乙方.*不得.*提前.*终止|甲方.*终止.*无须.*赔偿", "确保双方提前解约权利对等", "《民法典》第722条"],
        ["LEA_TER002", "终止风险", "续租优先权缺失", "warning", "无.*优先续租权|续租.*另行协商|优先续租.*未约定", "建议约定承租人优先续租权", "《民法典》第732条"],
        ["LEA_TER003", "终止风险", "解约赔偿不对等", "warning", "甲方.*违约.*赔偿.*\\d+%.+乙方.*违约.*赔偿", "双方解约赔偿条件应大致对等", ""],
        ["LEA_TER004", "终止风险", "提前解约通知期过长", "warning", "提前.*\\d+.*个月.*通知|终止.*须.*\\d+.*个月.*书面", "提前解约通知期应合理", ""],
        ["LEA_CON001", "物况风险", "交付标准不明确", "critical", "交付标准.*另行约定|交付.*状态.*未明确", "明确租赁物交付时的状态和标准", "《民法典》第708条"],
        ["LEA_CON002", "物况风险", "瑕疵担保责任免除", "warning", "不承担.*瑕疵.*责任|交付.*现状.*为准|瑕疵.*乙方.*自行.*负责", "出租人应对租赁物承担瑕疵担保责任", "《民法典》第731条"],
        ["LEA_CON003", "物况风险", "返还标准缺失", "info", "返还.*标准.*未约定|返还.*状态.*另行协商", "明确租赁物返还时的标准", "《民法典》第733条"],
    ]

    for row in data:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20

    wb.save(os.path.join(outputDir, "租赁合同风险规则.xlsx"))
    print(f"已创建: 租赁合同风险规则.xlsx")


if __name__ == "__main__":
    print("正在创建行业风险库 Excel 文件...")
    createBaseRisksExcel()
    createProcurementRisksExcel()
    createLeasingRisksExcel()
    print("\n所有 Excel 文件创建完成！")
    print(f"文件位置: {outputDir}")
